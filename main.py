import zipfile
import tarfile
import os
import rarfile
import py7zr
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from docx import Document
import PyPDF2
import hashlib
from PIL import Image, ImageTk
import io
import pandas as pd
import openpyxl

def resize_background(event):
    try:
        new_width = event.width
        new_height = event.height
        resized_image = bg_image_original.resize((new_width, new_height), Image.LANCZOS)
        bg_photo = ImageTk.PhotoImage(resized_image)
        canvas.image = bg_photo
    except Exception as e:
        messagebox.showerror("Error",f"Error resizing background: \n{e}")

def calculate_image_hash(image_bytes, algorithm='sha256'):
    hash_func = hashlib.new(algorithm)
    try:
        with Image.open(image_bytes) as img:
            buffered = io.BytesIO()
            img.save(buffered, format='PNG')
            image_data = buffered.getvalue()
        
        hash_func.update(image_data)
        return hash_func.hexdigest()

    except Exception as e:
        messagebox.showerror("Error",f"Error calculating hash: \n{e}")
        return "Error"


def extract_images_from_pdf(file_path):
    try:
        pdf_reader = PyPDF2.PdfReader(file_path)
        for page in pdf_reader.pages:
            xObject = page['/Resources']['/XObject'].get_object()
            if not xObject:
                continue
            for obj in xObject:
                if xObject[obj]['/Subtype'] == '/Image':
                    image_data = xObject[obj].get_data()
      
                    image_hash = calculate_image_hash(io.BytesIO(image_data))  
                    return image_hash
        return None
    except Exception as e:
        messagebox.showerror("Error",f"Error reading PDF file {file_path}: \n{e}")
        return "Error"

def extract_image_from_docx(file_path):
    try:
        document = Document(file_path)
        for rel in document.part.rels.values():
            if "image" in rel.target_ref:
                image_data = rel.target_part.blob  
     
                image_hash = calculate_image_hash(io.BytesIO(image_data))  
                return image_hash
        return None
    except Exception as e:
        messagebox.showerror("Error",f"Error extracting image from DOCX file {file_path}: \n{e}")
        return "Error"

def extract_zip(file_path, extract_path):
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(extract_path)
    except Exception as e:
        messagebox.showerror("Error",f"Error extracting ZIP file {file_path}: \n{e}")


def extract_tar(file_path, extract_path):
    try:
        with tarfile.open(file_path, 'r:*') as tar_ref:
            tar_ref.extractall(extract_path)
    except Exception as e:
        messagebox.showerror("Error",f"Error extracting TAR file {file_path}: \n{e}")


def extract_rar(file_path, extract_path):
    try:
        with rarfile.RarFile(file_path) as rar_ref:
            rar_ref.extractall(extract_path)
    except Exception as e:
        messagebox.showerror("Error",f"Error extracting RAR file {file_path}: \n{e}")


def extract_7z(file_path, extract_path):
    try:
        with py7zr.SevenZipFile(file_path, mode='r') as seven_zip_ref:
            seven_zip_ref.extractall(extract_path)
    except Exception as e:
        messagebox.showerror("Error",f"Error extracting 7Z file {file_path}: \n{e}")


def get_author_from_docx(file_path):
    try:
        document = Document(file_path)
        author = document.core_properties.author
        return author if author else "Unknown"
    except Exception as e:
        messagebox.showerror("Error",f"Error getting author from DOCX file {file_path}: \n{e}")
        return "Error"

def get_author_from_pdf(file_path):
    try:
        with open(file_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            author = pdf_reader.metadata.get('/Author')
            return author if author else "Unknown"
    except Exception as e:
        messagebox.showerror("Error",f"Error reading PDF {file_path}: \n{e}")
        return "Error"

def calculate_hash(file_path, algorithm='sha256'):
    hash_func = hashlib.new(algorithm)
    try:
        with open(file_path, 'rb') as file:
            while chunk := file.read(4096):  
                hash_func.update(chunk)
        return hash_func.hexdigest()
    except Exception as e:
        messagebox.showerror("Error",f"Error calculating hash for {file_path}: \n{e}")
        return "Error"


def browse_and_extract():
    for item in tree.get_children():
        tree.delete(item)
    folder_path = filedialog.askdirectory(title="Select a Folder")
    if folder_path:
        extract_path = os.path.join(folder_path, "Extracted Files")
        if not os.path.exists(extract_path):
            os.makedirs(extract_path)

        file_list = [f for f in os.listdir(folder_path) if f.endswith(('.zip', '.tar', '.tar.gz', '.tar.bz2', '.rar', '.7z'))]
        num_files = len(file_list)

        if num_files == 0:
            messagebox.showinfo("No Files", "No archive files found for extraction.")
            return


        progress = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
        progress.pack(pady=10)
        progress["maximum"] = num_files

        if extract_path:
            extracted_files = 0
            doc_author_info = []
            author_count = {}

            for filename in os.listdir(folder_path):
                file_path = os.path.join(folder_path, filename)
                try:
                    if filename.endswith('.zip'):
                        extract_zip(file_path, extract_path)
                        extracted_files += 1
                    elif filename.endswith(('.tar', '.tar.gz', '.tar.bz2')):
                        extract_tar(file_path, extract_path)
                        extracted_files += 1
                    elif filename.endswith('.rar'):
                        extract_rar(file_path, extract_path)
                        extracted_files += 1
                    elif filename.endswith('.7z'):
                        extract_7z(file_path, extract_path)
                        extracted_files += 1
                        
                except Exception as e:
                    messagebox.showerror("Error",f"Error processing {filename}: \n{e}")
                progress["value"] = extracted_files
                root.update_idletasks()
            for filename in os.listdir(extract_path):
                file_path = os.path.join(extract_path, filename)
                try:
                    if filename.endswith(('.doc', '.docx')):
                        image_hash = extract_image_from_docx(file_path)
                        author = get_author_from_docx(file_path)
                        file_hash = calculate_hash(file_path)
                        doc_author_info.append((filename, author, file_hash,image_hash))
                        author_count[author] = author_count.get(author, 0) + 1
                        author_count[image_hash] = author_count.get(image_hash, 0) + 1
                    elif filename.endswith('.pdf'):
                        image_hash = extract_images_from_pdf(file_path)
                        author = get_author_from_pdf(file_path)
                        file_hash = calculate_hash(file_path)
                        doc_author_info.append((filename, author, file_hash,image_hash))
                        author_count[author] = author_count.get(author, 0) + 1
                        author_count[image_hash] = author_count.get(image_hash, 0) + 1
                except Exception as e:
                    messagebox.showerror("Error",f"Error processing {filename}: \n{e}")
            if doc_author_info:
                tree.tag_configure('duplicate_author', background='red')
                tree.tag_configure('duplicate_image', background='yellow')
                tree.tag_configure('duplicate_both', background='green')

                for i, doc_info in enumerate(doc_author_info, start=1):
                    is_duplicate_author = author_count[doc_info[1]] > 1
                    is_duplicate_image = author_count[doc_info[3]] > 1 if doc_info[3] is not None else False

                    if is_duplicate_author and is_duplicate_image:
                        tag = 'duplicate_both'
                    elif is_duplicate_author:
                        tag = 'duplicate_author'
                    elif is_duplicate_image:
                        tag = 'duplicate_image'
                    else:
                        tag = ''  # No duplicate

                    tree.insert("", "end", values=(i,doc_info[0], doc_info[1], doc_info[2], doc_info[3]), tags=(tag,))
                
                progress.destroy()
        else:
            messagebox.showwarning("Warning", "No folder selected.")

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from tkinter import messagebox, filedialog

def export_tree_to_excel(tree):
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", "*.xlsx"),
                                                       ("All files", "*.*")],
                                             title="Save Treeview Data as Excel")
    if not file_path:
        return  # User canceled the save dialog

    try:
        data = []
        row_colors = []  
        
        for row_id in tree.get_children():
            row = tree.item(row_id)['values']
            row_tags = tree.item(row_id).get('tags', [])
            data.append(row)
            if 'duplicate_both' in row_tags:
                row_colors.append("90EE90")  
            elif 'duplicate_author' in row_tags:
                row_colors.append("FF6347")  
            elif 'duplicate_image' in row_tags:
                row_colors.append("FFFFE0")  
            else:
                row_colors.append("FFFFFF")  

        df = pd.DataFrame(data, columns=["Serial", "File Name", "Author", "File Hash", "Img Hash"])

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Extracted Data"

        headers = ["Serial", "File Name", "Author", "File Hash", "Img Hash"]
        worksheet.append(headers)

        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in worksheet[1]:
            cell.fill = header_fill

        for i, row_data in enumerate(data, start=2):
            row = worksheet.append(row_data)  
            row_fill = PatternFill(start_color=row_colors[i-2], end_color=row_colors[i-2], fill_type="solid")
            for cell in worksheet[i]:
                cell.fill = row_fill

        workbook.save(file_path)
        messagebox.showinfo("Success", "Treeview data exported successfully as Excel.")

    except Exception as e:
        messagebox.showerror("Error", f"Error exporting data: \n{e}")


def exitt():
    root.destroy()

# Tkinter Setup
root = tk.Tk()
root.title("MetaPlaig")
root.wm_state('zoomed')

image_path = r"C:\Prabhav\Python\Metaplaig\Images\backgroundd.jpg"  
try:
    bg_image_original = Image.open(image_path)
    bg_photo = ImageTk.PhotoImage(bg_image_original)

    canvas = tk.Canvas(root, highlightthickness=0)
    canvas.create_image(0, 0, anchor="nw", image=bg_photo)
    canvas.pack(fill="both", expand=True)
    root.bind("<Configure>", resize_background)
    
except Exception as e:
    messagebox.showerror("Error", f"Failed to load background image: {e}")
    
extract_button = tk.Button(root, text="Extract Files", command=browse_and_extract, font=("Arial", 12, "bold"),bg="yellow",fg="black",activebackground="#FFD700",activeforeground="black")
extract_button.pack()
export_excel_button = tk.Button(root, text="Export to Excel", command=lambda: export_tree_to_excel(tree), font=("Arial", 12, "bold"),bg="yellow",fg="black",activebackground="#FFD700",activeforeground="black")
export_excel_button.pack()
exit_button = tk.Button(root, text="Exit", command=exitt, font=("Arial", 12, "bold"),bg="red",fg="white",activebackground="darkred",activeforeground="white")
exit_button.pack()

canvas.create_window(470, 50, window=extract_button)
canvas.create_window(670, 50, window=export_excel_button)
canvas.create_window(570, 50, window=exit_button)


scrollbary = tk.Scrollbar(root, orient="vertical")
scrollbarx = tk.Scrollbar(root, orient="horizontal")
tree = ttk.Treeview(root, columns=("Serial", "File Name", "Author", "File Hash", "Img Hash"), show="headings",height=20)

tree.heading("Serial", text="Serial")
tree.heading("File Name", text="File Name")
tree.heading("Author", text="Author")
tree.heading("File Hash", text="File Hash")
tree.heading("Img Hash", text="Img Hash")
tree.column("Serial", width=50, anchor="center")
tree.column("File Name", width=300, anchor="w")
tree.column("Author", width=200, anchor="w")
tree.column("File Hash", width=300, anchor="w")
tree.column("Img Hash", width=300, anchor="w")

scrollbary.pack(side="right", fill="y")
scrollbarx.pack(side="bottom", fill="x")
tree.configure(yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)

scrollbary.configure(command=tree.yview)
scrollbarx.configure(command=tree.xview)

canvas.create_window(630, 300, window=tree)

scrollbary.pack_forget()
scrollbarx.pack_forget()

root.mainloop()
